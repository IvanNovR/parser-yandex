import openpyxl
import sys
from requests import get
from bs4 import BeautifulSoup
from PyQt5 import QtWidgets
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QWidget, QApplication
from design import Ui_MainWindow
from script import PostID

HEADERS = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.41 Safari/537.36 Edg/101.0.1210.32',
           'accept': '*/*'}
STATES = {
    "Киров": 43,
    "Марий-Эл": 12,
    "Мордовия": 13,
    "Нижний Новгород": 52,
    "Оренбург": 56,
    "Пенза": 58,
    "Самара": 63,
    "Саратов": 64,
    "Удмуртия": 18,
    "Ульяновск": 73,
    "Чувашия": 21,
    "Татарстан": 16
}
mrf = {
    'Киров': 'МРФ Волга / Кировский филиал',
    'Марий-Эл': 'МРФ Волга / Филиал в Республике Марий Эл',
    'Мордовия': 'МРФ Волга / Филиал в Республике Мордовия',
    'Нижний Новгород': 'МРФ Волга / Нижегородский филиал',
    'Оренбург': 'МРФ Волга / Оренбургский филиал',
    'Пенза': 'МРФ Волга / Пензенский филиал',
    'Самара': 'МРФ Волга / Самарский филиал',
    'Саратов': 'МРФ Волга / Саратовский филиал',
    'Удмуртия': 'МРФ Волга / Филиал в Удмуртской Республике',
    'Ульяновск': 'МРФ Волга / Ульяновский филиал',
    'Чувашия': 'МРФ Волга / Филиал в Чувашской Республике',
    'Татарстан': 'МРФ Волга / Филиал в республике Татарстан'
}


def parser(input_name, output_name, inn_column, filial_column):
    inn = openpyxl.load_workbook(input_name)
    innws = inn.active
    wb = openpyxl.load_workbook(output_name)
    ws = wb["License1"]

    k = 1
    while k < 2400:
        for i in range(4, 180):
            takeinn = innws[f"{inn_column}{i}"].value
            print(takeinn)
            takecity = innws[f"{filial_column}{i}"].value
            mycity = 0
            infocity = 'Филиал не найден'
            for key, value in zip(STATES.keys(), STATES.values()):
                if takecity == key:
                    mycity = value
                    break
            for key, value in zip(mrf.keys(), mrf.values()):
                if takecity == key:
                    infocity = value
                    break
            getid = PostID(org_inn=takeinn, service_id=3, region_id=int(mycity))
            a = getid.dopost()
            if a == '-':
                continue
            for number in range(len(a)):
                idlink = a[number].text
                r = get(f'https://rkn.gov.ru/communication/register/license/?id={idlink}', headers=HEADERS)
                html = BeautifulSoup(r.content, 'html.parser')
                table = html.select(".TblList")
                td = table[0].find_all('td')
                removetd = lambda x: x.removeprefix("<td>").removesuffix("</td>")
                INFO = {
                    "CITY": str(infocity),
                    "ID": removetd(str(td[3])),
                    "LICENSEDATE": removetd(str(td[5])),
                    "DATESTART": removetd(str(td[7])),
                    "VALIDUNTIL": removetd(str(td[11])),
                    "NAME": removetd(str(td[15])),
                    "INN": removetd(str(td[25])),
                    "DISTRICT": removetd(str(td[31]))
                }

                ws[f"A{k}"].value = INFO["CITY"]
                ws[f"B{k}"].value = INFO["NAME"]
                ws[f"C{k}"].value = INFO["INN"]
                ws[f"D{k}"].value = INFO["ID"]
                ws[f"F{k}"].value = INFO["DATESTART"]
                ws[f"E{k}"].value = INFO["LICENSEDATE"]
                ws[f"G{k}"].value = INFO["VALIDUNTIL"]
                ws[f"H{k}"].value = INFO["DISTRICT"]
                print(k, INFO)
                k += 1
                wb.save(output_name)
        break

class Photoredaktor(QtWidgets.QMainWindow, QWidget):
    def __init__(self):
        super(Photoredaktor, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.vSl = 100
        self.red = 0
        self.green = 0
        self.blue = 0
        self.flag = False
        self.init_Ui()


    def init_Ui(self):
        self.setMouseTracking(True)
        self.setWindowTitle('PythonParser')
        self.setWindowIcon(QIcon('icon.png'))
        self.ui.WayLineEdit.setPlaceholderText('укажите путь в формате C:\Program Files\...\...')
        self.ui.SaveLineEdit.setPlaceholderText('укажите путь в формате C:\Program Files\...\...')
        self.ui.InnLineEdit.setPlaceholderText('укажите букву столбца с ИНН')
        self.ui.FilialLineEdit.setPlaceholderText('укажите букву столбца с филиалом')
        self.ui.StartButton.clicked.connect(self.start)


    def start(self):
        input_name = self.ui.WayLineEdit.text()
        output_name = self.ui.SaveLineEdit.text()
        inn_column = self.ui.InnLineEdit.text()
        filial_column = self.ui.FilialLineEdit.text()
        parser(input_name, output_name, inn_column, filial_column)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    application = Photoredaktor()
    application.show()
    sys.exit(app.exec())